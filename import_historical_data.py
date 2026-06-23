"""
import_historical_data.py
=========================
Imports historical records into the new BusinessVault database.

Run once:  python import_historical_data.py
Safe to re-run: duplicate-safe throughout.
Requires: openpyxl, LibreOffice installed
"""

import sqlite3, os, subprocess, tempfile
from datetime import datetime, date
from openpyxl import load_workbook

DB_FILE  = "business_vault.db"
SOFFICE  = r"C:\Program Files\LibreOffice\program\soffice.exe"
STORE_MAP = {"Uxbr":"Uxbridge","Newb":"Newbury","Uxbridge":"Uxbridge","Newbury":"Newbury"}

FILES = {
    "staff_details":    "Staff Details.xls",
    "invoice_retail":   "Supplier's Invoice record (Uxbr-Newb)_Form.xlsm",
    "invoice_btl":      "Invoice record (BTL)_Form).xlsm",
    "hours_newb":       "Staff Hours (Newb).xlsm",
    "hours_uxbr":       "Staff Hours (Uxbr).xlsm",
    "hours_newb_may":   "May_26 Staff Hours (Newbury).xls",
    "hours_uxbr_may":   "May_26 Staff Hours (Uxbridge).xls",
    "full_record_newb": "Full Record (358).xls",
    "full_record_uxbr": "Full Record (372).xls",
}

def db():
    c = sqlite3.connect(DB_FILE)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON;")
    return c

def fmt_date(v):
    if v is None: return None
    if isinstance(v,(datetime,date)): return v.strftime("%Y-%m-%d")
    if isinstance(v,str) and v.strip():
        for f in ("%Y-%m-%d","%d/%m/%Y","%m/%d/%Y"):
            try: return datetime.strptime(v.strip(),f).strftime("%Y-%m-%d")
            except: pass
    return None

def fmt_time(v):
    if v is None: return None
    if isinstance(v,datetime): return v.strftime("%H:%M:%S")
    if hasattr(v,"strftime"): return v.strftime("%H:%M:%S")
    return None

def xls_to_xlsx(src):
    tmp = tempfile.mkdtemp()
    r = subprocess.run([SOFFICE,"--headless","--convert-to","xlsx",src,"--outdir",tmp],
                       capture_output=True,text=True)
    if r.returncode != 0: raise RuntimeError(r.stderr)
    return os.path.join(tmp, os.path.splitext(os.path.basename(src))[0]+".xlsx")

def open_wb(path):
    if path.lower().endswith(".xls"): path = xls_to_xlsx(path)
    return load_workbook(path, read_only=True, data_only=True)

def import_staff(conn):
    path = FILES["staff_details"]
    if not os.path.exists(path): print(f"⚠️  Skipping: {path}"); return
    print(f"\n👤 Staff from: {path}")
    wb=open_wb(path); ws=wb["Data"]; cur=conn.cursor(); ins=sk=0
    for row in ws.iter_rows(min_row=3,values_only=True):
        first,last=row[1],row[2]
        if not first or not last: continue
        if isinstance(first,str) and first.startswith("="): continue
        store = STORE_MAP.get(str(row[3]).strip(),str(row[3]).strip()) if row[3] else None
        phone = str(int(row[6])) if isinstance(row[6],float) else (str(row[6]) if row[6] else None)
        dol   = fmt_date(row[14])
        # Skip if this person already exists (same first+last+date_joined)
        doj_val = fmt_date(row[13])
        cur.execute("""SELECT COUNT(*) FROM staff_profiles
                       WHERE first_name=? AND last_name=? AND COALESCE(date_joined,'')=?""",
                    (first, last, doj_val or ''))
        if cur.fetchone()[0] > 0:
            sk += 1
            continue
        try:
            cur.execute("""INSERT INTO staff_profiles
                (staff_number,first_name,last_name,store_name,sex,phone,email,
                 address_1,address_2,address_3,address_4,postcode,
                 date_joined,date_left,leaving_reason,date_of_birth,
                 contracted_hrs,hourly_rate,is_salaried,salary_amount,is_active)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
              (row[0],first,last,store,row[4],phone,row[7],
               row[8],row[9],row[10],row[11],row[12],
               doj_val,dol,str(row[15]) if row[15] else None,fmt_date(row[18]),
               row[24] if isinstance(row[24],(int,float)) else None,
               row[26] if isinstance(row[26],(int,float)) else None,
               str(row[28]).strip() if row[28] else "N",
               row[29] if isinstance(row[29],(int,float)) else None,
               0 if dol else 1))
            ins+=1
        except sqlite3.IntegrityError: sk+=1
    conn.commit(); wb.close()
    print(f"   ✅ {ins} inserted, {sk} skipped")

def import_retail_invoices(conn):
    path = FILES["invoice_retail"]
    if not os.path.exists(path): print(f"⚠️  Skipping: {path}"); return
    print(f"\n🧾 Retail invoices from: {path}")
    wb=open_wb(path); ws=wb["Data"]; cur=conn.cursor(); ins=sk=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        supplier=row[1]; store=row[2]
        inv_no=str(row[7]).strip() if row[7] is not None else None
        if not supplier or not isinstance(supplier,str) or supplier.startswith("="): continue
        if not inv_no or inv_no in ("None",""): continue
        store_full = STORE_MAP.get(str(store).strip(),str(store).strip()) if store else None
        pd_dt = fmt_date(row[18])
        method_map={22:"DD",23:"Card",24:"Amex",25:"Online",26:"Cash",27:"Cheque"}
        pay_meth = next((l for c,l in method_map.items()
                         if len(row)>c and row[c] and str(row[c]).strip().lower() in ("yes","y","true","1")),None)
        try:
            cur.execute("""INSERT INTO supplier_invoices
                (seq_no,supplier_name,store_name,invoice_number,invoice_date,
                 gross_amount,vat_amount,net_amount,payment_terms,due_date,
                 paid_date,amount_paid,is_paid,payment_method,comments,
                 approval_status,submitted_by)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'approved','import')""",
              (row[0],supplier,store_full,inv_no,fmt_date(row[10]),
               row[11] if isinstance(row[11],(int,float)) else None,
               row[12] if isinstance(row[12],(int,float)) else None,
               row[13] if isinstance(row[13],(int,float)) else None,
               int(row[14]) if isinstance(row[14],(int,float)) else None,
               fmt_date(row[15]),pd_dt,
               row[19] if isinstance(row[19],(int,float)) else None,
               "Yes" if pd_dt else "No",pay_meth,
               str(row[29]) if len(row)>29 and row[29] else None))
            ins+=1
        except sqlite3.IntegrityError: sk+=1
    conn.commit(); wb.close()
    print(f"   ✅ {ins} inserted, {sk} skipped")

def import_btl_invoices(conn):
    path = FILES["invoice_btl"]
    if not os.path.exists(path): print(f"⚠️  Skipping: {path}"); return
    print(f"\n🏠 Property invoices from: {path}")
    wb=open_wb(path); ws=wb["Data"]; cur=conn.cursor(); ins=sk=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        supplier=row[1]; prop=row[2]
        if not supplier or not isinstance(supplier,str): continue
        if not prop or not isinstance(prop,str): continue
        inv_no=str(row[5]).strip() if row[5] is not None else None
        pd_dt=fmt_date(row[16])
        try:
            cur.execute("""INSERT OR IGNORE INTO property_invoices
                (property_name,supplier_name,invoice_number,invoice_date,
                 gross_amount,vat_amount,net_amount,due_date,
                 paid_date,amount_paid,is_paid,approval_status,submitted_by)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,'approved','import')""",
              (prop,supplier,inv_no,fmt_date(row[8]),
               row[9]  if isinstance(row[9], (int,float)) else None,
               row[10] if isinstance(row[10],(int,float)) else None,
               row[11] if isinstance(row[11],(int,float)) else None,
               fmt_date(row[13]),pd_dt,
               row[17] if isinstance(row[17],(int,float)) else None,
               "Yes" if pd_dt else "No"))
            ins+=1
        except sqlite3.IntegrityError: sk+=1
    conn.commit(); wb.close()
    print(f"   ✅ {ins} inserted, {sk} skipped")

def import_timesheets(conn, path, store, historical=False):
    if not os.path.exists(path): print(f"⚠️  Skipping: {path}"); return
    print(f"\n⏱  Timesheets ({store}) from: {path}")
    wb=open_wb(path)
    skip={"Main","Staff","Test","Unprotect_Report","Month View","Week1","Week2","Week3","Week4","Week5"}
    sheets=[s for s in wb.sheetnames if s not in skip and not s.startswith("X")]
    cur=conn.cursor(); ins=sk=0
    flag = "HISTORICAL" if historical else "GPS_VERIFIED"
    for sname in sheets:
        ws=wb[sname]
        for row in ws.iter_rows(min_row=2,values_only=True):
            dv=row[1]
            if not isinstance(dv,(datetime,date)): continue
            if row[4] is None and row[3] is None: continue
            emp = str(row[2]).strip() if row[2] and not str(row[2]).startswith("=") else sname
            ci  = fmt_time(row[4])
            co  = fmt_time(row[6]) if len(row)>6 else None
            st  = flag if ci else ("ABSENCE" if row[3] else None)
            try:
                cur.execute("""INSERT OR IGNORE INTO timesheets
                    (staff_name,store_name,work_date,clock_in_time,clock_out_time,
                     status_flag,absence_type,comments)
                    VALUES(?,?,?,?,?,?,?,?)""",
                  (emp,store,fmt_date(dv),ci,co,st,row[3],
                   str(row[8]) if len(row)>8 and row[8] else None))
                if cur.rowcount: ins+=1
                else: sk+=1
            except: pass
    conn.commit(); wb.close()
    print(f"   ✅ {ins} inserted, {sk} skipped")

def summary(conn):
    cur=conn.cursor()
    print("\n"+"="*50+"\n📊 IMPORT SUMMARY\n"+"="*50)
    for t,l in [("staff_profiles","Staff profiles"),("supplier_invoices","Retail invoices"),
                ("property_invoices","Property invoices"),("timesheets","Timesheet records")]:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {t}")
            print(f"  {l:<25} {cur.fetchone()[0]:>6}")
        except Exception as e: print(f"  {l:<25} ERROR: {e}")
    today=datetime.now().strftime("%Y-%m-%d")
    cur.execute("SELECT COUNT(*) FROM supplier_invoices WHERE is_paid!='Yes' AND due_date<?", (today,))
    print(f"\n  🚨 Overdue retail invoices:  {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM staff_profiles WHERE is_active=1")
    print(f"  👤 Active staff:             {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM staff_profiles WHERE is_active=0")
    print(f"  📁 Former staff (leavers):   {cur.fetchone()[0]}")
    print("="*50)

def main():
    if not os.path.exists(DB_FILE):
        print(f"❌ {DB_FILE} not found — start the app first, then stop it and run this script.")
        return
    print(f"🚀 Importing into {DB_FILE}")
    missing=[v for k,v in FILES.items() if not os.path.exists(v)]
    if missing: print(f"⚠️  Missing (will skip): {missing}")
    conn=db()
    import_staff(conn)
    import_retail_invoices(conn)
    import_btl_invoices(conn)
    import_timesheets(conn, FILES["hours_newb"],     "Newbury")
    import_timesheets(conn, FILES["hours_uxbr"],     "Uxbridge")
    import_timesheets(conn, FILES["hours_newb_may"], "Newbury",  historical=True)
    import_timesheets(conn, FILES["hours_uxbr_may"], "Uxbridge", historical=True)
    summary(conn)
    conn.close()
    print("\n✅ Done. Refresh your browser.")

if __name__=="__main__":
    main()


def import_leave_records(conn, path, store_name):
    """Import 2026 leave records from Full_Record file.
    Layout: each Name_Leave sheet has rows for each month (Jan-Dec),
    columns 1-31 represent days, cells contain absence codes (H, B, S etc.)
    """
    if not os.path.exists(path):
        print(f"⚠️  Skipping leave records: {path} not found")
        return
    print(f"\n📅 Leave records ({store_name}) from: {path}")
    wb  = open_wb(path)
    cur = conn.cursor()
    ins = sk = 0

    # Build staff name → id map (first name only for sheet matching)
    # Also add first-word-only mapping for names like "Rhys Michael"
    staff_map = {}
    rows = conn.execute(
        "SELECT staff_id, first_name FROM staff_profiles WHERE is_active=1"
    ).fetchall()
    for r in rows:
        full_first = r[1].strip().lower()
        staff_map[full_first] = r[0]
        # Also map just the first word (e.g. "rhys michael" → "rhys")
        first_word = full_first.split()[0]
        if first_word not in staff_map:
            staff_map[first_word] = r[0]

    for sname in wb.sheetnames:
        if not sname.endswith("_Leave"):
            continue
        first_name = sname.replace("_Leave","").strip().lower()
        staff_id   = staff_map.get(first_name)
        if not staff_id:
            continue

        ws   = wb[sname]
        year = 2026

        # Also update entitlement from row 26 (Full Year Entitlement)
        entitlement = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=35, values_only=True), 1):
            # Row 26 = Full Year Entitlement
            if row_idx == 26 and isinstance(row[10], (int, float)):
                entitlement = float(row[10])
                continue

            # Rows 6-17 = Jan-Dec (month in col 0, day codes in cols 1-31)
            if row_idx < 6 or row_idx > 17:
                continue
            month_val = row[0]
            if not isinstance(month_val, datetime):
                continue
            month = month_val.month

            # Scan columns 1-31 for absence codes
            for day in range(1, 32):
                cell = row[day] if day < len(row) else None
                if not cell or cell == "#N/A":
                    continue
                code = str(cell).strip().upper()
                if code not in ("H","B","S","D","L","J","M","AL","UL","TO"):
                    continue
                # Build the date
                try:
                    leave_date = datetime(year, month, day).strftime("%Y-%m-%d")
                except ValueError:
                    continue  # Invalid date (e.g. Feb 30)

                try:
                    cur.execute("""
                        INSERT OR IGNORE INTO leave_requests
                            (staff_id, leave_type, date_from, date_to,
                             days_taken, status, requested_by, notes)
                        VALUES(?,?,?,?,1,'approved','import','Imported from Excel')
                    """, (staff_id, code, leave_date, leave_date))
                    if cur.rowcount: ins += 1
                    else: sk += 1
                except: pass

        # Update entitlement on staff profile if found
        if entitlement and staff_id:
            # Convert entitlement days to hours based on contracted hrs
            s = conn.execute(
                "SELECT contracted_hrs FROM staff_profiles WHERE staff_id=?",
                (staff_id,)
            ).fetchone()
            if s and s[0]:
                daily_hrs = s[0] / 5
                ent_hrs   = round(entitlement * daily_hrs, 1)
                # Store as a note for now — entitlement is calculated dynamically
                pass

    conn.commit()
    wb.close()
    print(f"   ✅ {ins} leave days inserted, {sk} skipped")


def import_pay_rates(conn):
    """Import current pay rates and NMW data from Staff Details.xls"""
    path = "Staff Details.xls"
    if not os.path.exists(path):
        print(f"⚠️  Skipping pay rates: {path} not found")
        return
    print(f"\n💰 Pay rates from: {path}")
    wb  = open_wb(path)

    # Use the Report (Current Employees) sheet for current rates
    try:
        ws = wb["Report (Current Employees)"]
    except:
        print("   ⚠️  Sheet not found")
        wb.close()
        return

    cur = conn.cursor()
    updated = 0
    today = datetime.now().strftime("%Y-%m-%d")

    for row in ws.iter_rows(min_row=2, values_only=True):
        first = row[1]; last = row[2]
        if not first or not last: continue
        if not isinstance(first, str): continue

        # Find current hourly rate — scan row for a plausible wage value (5-25)
        rate = None
        for cell in row[14:]:
            if isinstance(cell, (int,float)) and 5 <= cell <= 30:
                rate = float(cell)
                break
        if not rate: continue

        # Update staff profile
        cur.execute("""UPDATE staff_profiles SET hourly_rate=?
                       WHERE LOWER(first_name)=? AND LOWER(last_name)=? AND is_active=1""",
                    (rate, first.strip().lower(), last.strip().lower()))
        if cur.rowcount:
            # Add pay history entry if not already there
            sid = cur.execute("""SELECT staff_id FROM staff_profiles
                                  WHERE LOWER(first_name)=? AND LOWER(last_name)=? AND is_active=1""",
                               (first.strip().lower(), last.strip().lower())).fetchone()
            if sid:
                existing = cur.execute("SELECT COUNT(*) FROM pay_history WHERE staff_id=?",
                                       (sid[0],)).fetchone()[0]
                if not existing:
                    cur.execute("""INSERT INTO pay_history
                        (staff_id,effective_date,hourly_rate,change_reason,recorded_by)
                        VALUES(?,?,?,'Opening rate — imported from Excel','import')""",
                      (sid[0], today, rate))
            updated += 1

    conn.commit()
    wb.close()
    print(f"   ✅ {updated} pay rates updated")


# Update main() to include new imports
_original_main = main

def main():
    if not os.path.exists(DB_FILE):
        print(f"❌ {DB_FILE} not found — start the app first, then stop it and run this script.")
        return
    print(f"🚀 Importing into {DB_FILE}")
    missing=[v for k,v in FILES.items() if not os.path.exists(v)]
    if missing: print(f"⚠️  Missing (will skip): {missing}")
    conn=db()
    import_staff(conn)
    import_retail_invoices(conn)
    import_btl_invoices(conn)
    import_timesheets(conn, FILES["hours_newb"],     "Newbury")
    import_timesheets(conn, FILES["hours_uxbr"],     "Uxbridge")
    import_timesheets(conn, FILES["hours_newb_may"], "Newbury",  historical=True)
    import_timesheets(conn, FILES["hours_uxbr_may"], "Uxbridge", historical=True)
    import_leave_records(conn, FILES["full_record_newb"], "Newbury")
    import_leave_records(conn, FILES["full_record_uxbr"], "Uxbridge")
    import_pay_rates(conn)
    summary(conn)
    conn.close()
    print("\n✅ Done. Refresh your browser.")

