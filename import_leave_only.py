"""
import_leave_only.py
====================
Imports ONLY leave records from Full Record files.
Safe to run multiple times.
"""
import sqlite3, subprocess, tempfile, os
from openpyxl import load_workbook
from datetime import datetime

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"
DB_FILE = "business_vault.db"

LEAVE_FILES = {
    "Full Record (358).xls": "Newbury",
    "Full Record (372).xls": "Uxbridge",
}

VALID_CODES = {"H","B","S","D","L","J","M","AL","UL","TO"}

def xls_to_xlsx(src):
    tmp = tempfile.mkdtemp()
    subprocess.run([SOFFICE,"--headless","--convert-to","xlsx",src,"--outdir",tmp],
                   capture_output=True, text=True)
    return os.path.join(tmp, os.path.splitext(os.path.basename(src))[0]+".xlsx")

def main():
    conn = sqlite3.connect(DB_FILE)
    cur  = conn.cursor()

    # Build staff name map — first name AND first word of first name
    staff_map = {}
    for r in conn.execute("SELECT staff_id, first_name FROM staff_profiles WHERE is_active=1").fetchall():
        full = r[1].strip().lower()
        staff_map[full] = r[0]
        first_word = full.split()[0]
        if first_word not in staff_map:
            staff_map[first_word] = r[0]

    print("Staff map:", {k: v for k,v in staff_map.items()})

    total_ins = total_sk = 0

    for fname, store in LEAVE_FILES.items():
        if not os.path.exists(fname):
            print(f"⚠️  Not found: {fname}")
            continue

        print(f"\n📅 {fname} ({store})")
        xlsx = xls_to_xlsx(fname)
        wb   = load_workbook(xlsx, read_only=True, data_only=True)
        ins  = sk = 0

        for sname in wb.sheetnames:
            if not sname.endswith("_Leave"):
                continue
            first = sname.replace("_Leave","").strip().lower()
            staff_id = staff_map.get(first)
            if not staff_id:
                print(f"  ⚠️  No match for sheet '{sname}' (first='{first}')")
                continue

            ws   = wb[sname]
            year = 2026

            for row_idx, row in enumerate(ws.iter_rows(max_row=18, values_only=True), 1):
                if row_idx < 6:
                    continue
                month_val = row[0]
                if not isinstance(month_val, datetime):
                    continue
                month = month_val.month

                for day in range(1, 32):
                    cell = row[day] if day < len(row) else None
                    if not cell or str(cell) == "#N/A":
                        continue
                    code = str(cell).strip().upper()
                    if code not in VALID_CODES:
                        continue
                    try:
                        leave_date = datetime(year, month, day).strftime("%Y-%m-%d")
                    except ValueError:
                        continue

                    try:
                        cur.execute("""
                            INSERT OR IGNORE INTO leave_requests
                                (staff_id, leave_type, date_from, date_to,
                                 days_taken, status, requested_by, notes)
                            VALUES(?,?,?,?,1,'approved','import','Imported from Excel')
                        """, (staff_id, code, leave_date, leave_date))
                        if cur.rowcount:
                            ins += 1
                        else:
                            sk += 1
                    except Exception as e:
                        print(f"  Error: {e}")

        conn.commit()
        wb.close()
        print(f"  ✅ {ins} leave days inserted, {sk} skipped")
        total_ins += ins
        total_sk  += sk

    print(f"\n{'='*40}")
    print(f"Total: {total_ins} inserted, {total_sk} skipped")

    # Show leave summary per staff
    print("\nLeave records per staff member:")
    for r in conn.execute("""
        SELECT sp.first_name, sp.last_name, lr.leave_type, COUNT(*) as n
        FROM leave_requests lr
        JOIN staff_profiles sp ON lr.staff_id=sp.staff_id
        GROUP BY lr.staff_id, lr.leave_type
        ORDER BY sp.first_name
    """).fetchall():
        print(f"  {r[0]} {r[1]}: {r[2]} × {r[3]} days")

    conn.close()

if __name__ == "__main__":
    main()
