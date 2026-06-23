import sqlite3, subprocess, tempfile, os
from openpyxl import load_workbook
from datetime import datetime

SOFFICE = r"C:\Program Files\LibreOffice\program\soffice.exe"

def xls_to_xlsx(src):
    tmp = tempfile.mkdtemp()
    subprocess.run([SOFFICE,"--headless","--convert-to","xlsx",src,"--outdir",tmp],
                   capture_output=True, text=True)
    return os.path.join(tmp, os.path.splitext(os.path.basename(src))[0]+".xlsx")

conn = sqlite3.connect("business_vault.db")

# Show active staff names in DB
print("=== Active staff in database ===")
staff_map = {}
for r in conn.execute("SELECT staff_id, first_name FROM staff_profiles WHERE is_active=1").fetchall():
    staff_map[r[1].strip().lower()] = r[0]
    print(f"  '{r[1]}' → id {r[0]}")

# Show _Leave sheets in Full Record file
print()
for fname in ["Full Record (358).xls", "Full Record (372).xls"]:
    if not os.path.exists(fname):
        print(f"NOT FOUND: {fname}")
        continue
    print(f"\n=== {fname} — Leave sheets ===")
    xlsx = xls_to_xlsx(fname)
    wb   = load_workbook(xlsx, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        if not sname.endswith("_Leave"):
            continue
        first = sname.replace("_Leave","").strip().lower()
        match = staff_map.get(first)
        print(f"  Sheet '{sname}' → first='{first}' → DB match: {match}")

        # Count non-empty leave cells
        ws    = wb[sname]
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(max_row=18, values_only=True), 1):
            if row_idx < 6: continue
            for day in range(1, 32):
                cell = row[day] if day < len(row) else None
                if cell and str(cell).strip().upper() in ("H","B","S","D","L","J","M"):
                    count += 1
        print(f"    → {count} leave codes found in sheet")
    wb.close()

conn.close()
